"""
Consolidate all eight validated metrics — revenue, gross margin,
operating margin, ROIC, customer concentration (the "moat" section)
plus debt/EBITDA, net debt/EBITDA, and interest coverage (the
"credit/leverage" section) — for Vertiv (VRT), Eaton (ETN), Quanta
Services (PWR), and nVent (NVT) into a single comparison table.

Every number in this table comes from calling the already-validated
functions in fetch_margins.py, fetch_roic_others.py,
fetch_customer_concentration.py, and fetch_credit_metrics.py directly:
- Revenue, gross margin, operating margin: fetch_margins.compute_margins()
- ROIC: fetch_roic_others.compute_roic(), run against the exact same
  companyfacts response already fetched for margins — not re-fetched,
  not re-derived.
- Customer concentration: fetch_customer_concentration's own
  most_recent_10k() / fetch_10k_text() / customer_concentration_sentences()
  pull the filing text and extract candidate sentences exactly as that
  script does standalone. The only new code here is
  concentration_headline(), a thin selection/formatting layer that
  picks which already-extracted sentence is the table's headline
  figure (it does not scan the filing itself or invent a number).
- Debt/EBITDA, net debt/EBITDA, interest coverage:
  fetch_credit_metrics.compute_credit_metrics(), which itself reuses
  fetch_roic_others.total_debt_for_roic() and fetch_margins's
  operating income — not re-derived here either.

Each underlying compute_*() function still prints its own diagnostic
output (tag matches, proxy derivations, flags) as it runs, exactly as
when those scripts are run standalone, so every number in the table
can be traced back to its source. Note this means compute_margins()
runs twice per company in this script's output (once directly, once
inside compute_credit_metrics()) — same deterministic inputs both
times, just a duplicated diagnostic printout, not a second derivation.

"Peer average" for the spread columns is the simple mean across all
four companies for that metric (including the company itself). Spread
columns only apply to the moat-section numeric metrics (gross margin,
operating margin, ROIC), per the original design — the credit/leverage
columns are shown as plain values, not spreads.

SEC EDGAR requires a descriptive User-Agent on every request and asks
that callers stay under 10 requests/second; the underlying scripts
already enforce this via their own sec_get()/sleep().

Output:
- output/consolidated_comparison.xlsx
- output/consolidated_comparison.html
"""

import os
import re

import fetch_credit_metrics
import fetch_customer_concentration
import fetch_margins
import fetch_roic_others
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TICKERS = ["VRT", "ETN", "PWR", "NVT"]
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")

# --- Customer concentration headline selection -----------------------------
# Operates only on sentences already extracted by
# fetch_customer_concentration.customer_concentration_sentences(); does not
# re-scan the filing or invent a figure.

LARGEST_CUSTOMER_PATTERN = re.compile(
    r"largest customers?\s+(?:accounted for|represented)\s+(?:approximately\s+)?(\d+(?:\.\d+)?)\s*%",
    re.IGNORECASE,
)
TOP_N_CUSTOMERS_PATTERN = re.compile(
    r"(?:ten|top\s*10|top\s*ten)\s+largest customers\s+(?:accounted for|represented)\s+(?:approximately\s+)?(\d+(?:\.\d+)?)\s*%",
    re.IGNORECASE,
)
NO_CUSTOMER_PATTERN = re.compile(
    r"no (?:single )?customer[s]?\b.{0,80}?(\d+(?:\.\d+)?)\s*%", re.IGNORECASE
)


def concentration_headline(sentences):
    """Pick the headline customer-concentration figure and its source
    sentence from the list of sentences fetch_customer_concentration.py
    already flagged as relevant. Priority:
      1. A "largest customer(s) accounted for/represented X%" statement
         (plus a companion top-10/top-ten figure if the same sentence
         discloses one).
      2. An explicit "no customer(s) ... X%" statement.
      3. Nothing matched -> "Not disclosed".
    """
    for s in sentences:
        m = LARGEST_CUSTOMER_PATTERN.search(s)
        if m:
            headline = f"Largest customer: {m.group(1)}%"
            m2 = TOP_N_CUSTOMERS_PATTERN.search(s)
            if m2:
                headline += f"; top 10: {m2.group(1)}%"
            return headline, s

    for s in sentences:
        m = NO_CUSTOMER_PATTERN.search(s)
        if m:
            return f"No customer exceeds {m.group(1)}%", s

    return "Not disclosed", None


def fmt_pct(x):
    return f"{x:.2%}"


def fmt_spread(x):
    sign = "+" if x >= 0 else ""
    return f"{sign}{x * 100:.1f}pp"


def gather_rows():
    tickers_json = fetch_margins.sec_get("https://www.sec.gov/files/company_tickers.json").json()
    ciks = {}
    for ticker in TICKERS:
        match = next(v for v in tickers_json.values() if v["ticker"] == ticker)
        ciks[ticker] = str(match["cik_str"]).zfill(10)

    rows = []
    for ticker in TICKERS:
        cik = ciks[ticker]

        facts = fetch_margins.sec_get(f"https://data.sec.gov/api/xbrl/companyfacts/CIK{cik}.json").json()
        us_gaap = facts["facts"]["us-gaap"]

        margins = fetch_margins.compute_margins(ticker, us_gaap)
        roic = fetch_roic_others.compute_roic(ticker, us_gaap)
        credit = fetch_credit_metrics.compute_credit_metrics(ticker, us_gaap)

        filing = fetch_customer_concentration.most_recent_10k(cik)
        text, source_url = fetch_customer_concentration.fetch_10k_text(cik, filing["accession"], filing["document"])
        sentences = fetch_customer_concentration.customer_concentration_sentences(text)
        headline, source_sentence = concentration_headline(sentences)

        rows.append({
            "ticker": ticker,
            "fiscal_year": margins["fiscal_year"],
            "revenue": margins["revenue"],
            "gross_margin": margins["gross_margin"],
            "operating_margin": margins["operating_margin"],
            "roic": roic["roic"],
            "concentration_headline": headline,
            "concentration_source": source_sentence,
            "concentration_filing_url": source_url,
            "debt_ebitda": credit["leverage"],
            "net_debt_ebitda": credit["net_leverage"],
            "interest_coverage": credit["interest_coverage"],
            "credit_flags": credit["flags"],
        })

    return rows


def collect_credit_flags(rows):
    """Flatten each company's credit-metric flags into (ticker, flag) pairs,
    in table order, for footnoting — these are the same flags
    fetch_credit_metrics.py prints standalone, not new commentary."""
    flags = []
    for r in rows:
        for flag in r["credit_flags"]:
            flags.append((r["ticker"], flag))
    return flags


def add_spreads(rows):
    for metric in ("gross_margin", "operating_margin", "roic"):
        avg = sum(r[metric] for r in rows) / len(rows)
        for r in rows:
            r[f"{metric}_peer_avg"] = avg
            r[f"{metric}_spread"] = r[metric] - avg
    return rows


# Each column is (key, label, section). "section" groups columns for the
# visual divider between moat metrics and credit/leverage metrics:
# "id" = row identifiers (ticker, FY), not part of either group.
SECTION_LABELS = {"id": "", "moat": "MOAT METRICS", "credit": "CREDIT / LEVERAGE"}

COLUMNS = [
    ("ticker", "Ticker", "id"),
    ("fiscal_year", "FY", "id"),
    ("revenue", "Revenue", "moat"),
    ("gross_margin", "Gross Margin", "moat"),
    ("gross_margin_spread", "Gross Margin Spread vs. Peer Avg", "moat"),
    ("operating_margin", "Operating Margin", "moat"),
    ("operating_margin_spread", "Operating Margin Spread vs. Peer Avg", "moat"),
    ("roic", "ROIC", "moat"),
    ("roic_spread", "ROIC Spread vs. Peer Avg", "moat"),
    ("concentration_headline", "Customer Concentration", "moat"),
    ("debt_ebitda", "Debt / EBITDA", "credit"),
    ("net_debt_ebitda", "Net Debt / EBITDA", "credit"),
    ("interest_coverage", "Interest Coverage (EBITDA / Interest)", "credit"),
]


def column_groups():
    """Consecutive runs of columns sharing the same section, as
    (section, start_idx, end_idx_inclusive), 0-indexed into COLUMNS."""
    sections = [section for _, _, section in COLUMNS]
    groups = []
    start = 0
    for i in range(1, len(sections) + 1):
        if i == len(sections) or sections[i] != sections[start]:
            groups.append((sections[start], start, i - 1))
            start = i
    return groups


def formatted_cell(key, row):
    val = row[key]
    if key == "revenue":
        return f"${val:,.0f}"
    if key in ("gross_margin", "operating_margin", "roic"):
        return fmt_pct(val)
    if key.endswith("_spread"):
        return fmt_spread(val)
    if key in ("debt_ebitda", "net_debt_ebitda", "interest_coverage"):
        return f"{val:.2f}x"
    return val


def section_banner(headers, widths):
    """A header row that visually merges consecutive same-section columns
    into one centered label (e.g. 'CREDIT / LEVERAGE' spanning its three
    columns), built on a blank canvas aligned to the same column offsets
    fmt_row() uses, so it lines up exactly with the header/data rows below."""
    sep = " | "
    offsets = []
    pos = 0
    for w in widths:
        offsets.append(pos)
        pos += w + len(sep)
    total_len = pos - len(sep)

    canvas = list(" " * total_len)
    for section, s, e in column_groups():
        label = SECTION_LABELS.get(section, "")
        if not label:
            continue
        span_start = offsets[s]
        span_end = offsets[e] + widths[e]
        centered = label.center(span_end - span_start)
        canvas[span_start:span_end] = list(centered)
    return "".join(canvas)


def print_table(rows):
    headers = [label for _, label, _ in COLUMNS]
    table_rows = [[str(formatted_cell(key, r)) for key, _, _ in COLUMNS] for r in rows]
    widths = [max(len(h), max(len(row[i]) for row in table_rows)) for i, h in enumerate(headers)]

    def fmt_row(cells):
        return " | ".join(cell.ljust(widths[i]) for i, cell in enumerate(cells))

    print("=== Consolidated comparison: VRT / ETN / PWR / NVT (FY2025) ===")
    print(section_banner(headers, widths))
    print(fmt_row(headers))
    print("-+-".join("-" * w for w in widths))
    for row in table_rows:
        print(fmt_row(row))


def print_credit_flags(credit_flags):
    if not credit_flags:
        return
    print()
    print("Credit/leverage flags (hand-check against the filings):")
    for ticker, flag in credit_flags:
        print(f"  [{ticker}] {flag}")


SECTION_FILL_COLORS = {"id": "F2F2F2", "moat": "DCE6F1", "credit": "FCE4D6"}


def write_xlsx(rows, path, credit_flags):
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"

    headers = [label for _, label, _ in COLUMNS]

    # Row 1: section banner — merged, shaded blocks grouping the moat
    # columns and the credit/leverage columns.
    for section, s, e in column_groups():
        fill = PatternFill(start_color=SECTION_FILL_COLORS[section], end_color=SECTION_FILL_COLORS[section], fill_type="solid")
        for col in range(s + 1, e + 2):
            cell = ws.cell(row=1, column=col)
            cell.fill = fill
        if e > s:
            ws.merge_cells(start_row=1, start_column=s + 1, end_row=1, end_column=e + 1)
        label_cell = ws.cell(row=1, column=s + 1)
        label_cell.value = SECTION_LABELS[section]
        label_cell.font = Font(bold=True, color="1A1A1A")
        label_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2: column headers (existing dark-blue style).
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = label
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Rows 3..: data
    for row_offset, r in enumerate(rows):
        row_idx = 3 + row_offset
        for col_idx, (key, _, _) in enumerate(COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx).value = r[key]

    pct_cols = [i for i, (key, _, _) in enumerate(COLUMNS, start=1) if key in ("gross_margin", "operating_margin", "roic")]
    spread_cols = [i for i, (key, _, _) in enumerate(COLUMNS, start=1) if key.endswith("_spread")]
    multiple_cols = [i for i, (key, _, _) in enumerate(COLUMNS, start=1) if key in ("debt_ebitda", "net_debt_ebitda", "interest_coverage")]
    revenue_col = next(i for i, (key, _, _) in enumerate(COLUMNS, start=1) if key == "revenue")

    for row_idx in range(3, 3 + len(rows)):
        ws.cell(row=row_idx, column=revenue_col).number_format = "$#,##0"
        for col in pct_cols:
            ws.cell(row=row_idx, column=col).number_format = "0.00%"
        for col in multiple_cols:
            ws.cell(row=row_idx, column=col).number_format = "0.00\"x\""
        for col in spread_cols:
            cell = ws.cell(row=row_idx, column=col)
            cell.number_format = "+0.0%;-0.0%"
            if cell.value is not None:
                if cell.value > 0:
                    cell.font = Font(color="006100")
                elif cell.value < 0:
                    cell.font = Font(color="9C0006")

    for col_idx, (key, _, _) in enumerate(COLUMNS, start=1):
        max_len = max(
            [len(headers[col_idx - 1])] + [len(str(formatted_cell(key, r))) for r in rows]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, max_len + 2)

    ws.freeze_panes = "A3"

    if credit_flags:
        flag_row = 3 + len(rows) + 2
        heading_cell = ws.cell(row=flag_row, column=1)
        heading_cell.value = "Credit / Leverage — flags to hand-check"
        heading_cell.font = Font(bold=True)
        for i, (ticker, flag) in enumerate(credit_flags, start=1):
            cell = ws.cell(row=flag_row + i, column=1)
            cell.value = f"[{ticker}] {flag}"
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=flag_row + i, start_column=1, end_row=flag_row + i, end_column=len(COLUMNS))
            ws.row_dimensions[flag_row + i].height = 30

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)


HTML_TEMPLATE = """<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>Consolidated Metrics Comparison — VRT / ETN / PWR / NVT</title>
<style>
  body {{ font-family: -apple-system, Segoe UI, Helvetica, Arial, sans-serif; margin: 2rem; color: #1a1a1a; background: #fff; }}
  h1 {{ font-size: 1.3rem; margin-bottom: 0.2rem; }}
  .subtitle {{ color: #555; margin-bottom: 1.5rem; font-size: 0.9rem; }}
  table {{ border-collapse: collapse; width: 100%; font-size: 0.9rem; }}
  th, td {{ border: 1px solid #ddd; padding: 8px 10px; text-align: right; }}
  th {{ background: #1f4e78; color: #fff; text-align: center; }}
  th.section {{ font-size: 0.85rem; letter-spacing: 0.04em; }}
  th.section-moat {{ background: #dce6f1; color: #1a1a1a; }}
  th.section-credit {{ background: #fce4d6; color: #1a1a1a; }}
  th.section-id {{ background: #f2f2f2; color: #1a1a1a; }}
  td.divider-left, th.divider-left {{ border-left: 3px solid #999; }}
  td:first-child, th:first-child {{ text-align: left; font-weight: 600; }}
  td.concentration {{ text-align: left; }}
  tr:nth-child(even) td {{ background: #f7f9fb; }}
  .pos {{ color: #0a6b0a; }}
  .neg {{ color: #b3261e; }}
  .footnote {{ margin-top: 1.5rem; font-size: 0.8rem; color: #666; max-width: 800px; }}
  .flags {{ margin-top: 1rem; font-size: 0.85rem; color: #333; max-width: 800px; }}
  .flags h2 {{ font-size: 0.95rem; margin-bottom: 0.4rem; }}
  .flags li {{ margin-bottom: 0.4rem; }}
</style>
</head>
<body>
<h1>Consolidated Metrics Comparison</h1>
<div class="subtitle">VRT, ETN, PWR, NVT — FY2025 — spreads are vs. the peer average (mean of all four companies)</div>
<table>
<thead>
<tr>{section_row}</tr>
<tr>{header_row}</tr>
</thead>
<tbody>
{body_rows}
</tbody>
</table>
<div class="footnote">
Revenue, gross margin, and operating margin from fetch_margins.py; ROIC from fetch_roic_others.py;
customer concentration extracted from each company's most recent 10-K by fetch_customer_concentration.py;
debt/EBITDA, net debt/EBITDA, and interest coverage from fetch_credit_metrics.py (which itself reuses
total_debt_for_roic() and fetch_margins's operating income, not re-derived).
ETN operating margin uses the derived proxy (Revenue - COGS - SG&amp;A - R&amp;D); no tagged OperatingIncomeLoss for FY2025.
</div>
{flags_block}
</body>
</html>
"""


def write_html(rows, path, credit_flags):
    section_cells = []
    for section, s, e in column_groups():
        span = e - s + 1
        label = SECTION_LABELS.get(section, "")
        divider = " divider-left" if section == "credit" else ""
        colspan_attr = f' colspan="{span}"' if span > 1 else ""
        section_cells.append(f'<th class="section section-{section}{divider}"{colspan_attr}>{label}</th>')
    section_row = "".join(section_cells)

    credit_start = next(s for section, s, _ in column_groups() if section == "credit")
    divider_attr = ' class="divider-left"'
    header_row = "".join(
        f'<th{divider_attr if i == credit_start else ""}>{label}</th>'
        for i, (_, label, _) in enumerate(COLUMNS)
    )

    body_rows = []
    for r in rows:
        cells = []
        for i, (key, _, _) in enumerate(COLUMNS):
            val = formatted_cell(key, r)
            css_classes = []
            if i == credit_start:
                css_classes.append("divider-left")
            if key.endswith("_spread"):
                css_classes.append("pos" if r[key] >= 0 else "neg")
            if key == "concentration_headline":
                css_classes.append("concentration")
            css_attr = f' class="{" ".join(css_classes)}"' if css_classes else ""
            cells.append(f"<td{css_attr}>{val}</td>")
        body_rows.append(f"<tr>{''.join(cells)}</tr>")

    flags_block = ""
    if credit_flags:
        items = "\n".join(f"<li><strong>{ticker}</strong> — {flag}</li>" for ticker, flag in credit_flags)
        flags_block = f'<div class="flags"><h2>Credit / Leverage — flags to hand-check</h2><ul>{items}</ul></div>'

    html = HTML_TEMPLATE.format(
        section_row=section_row, header_row=header_row, body_rows="\n".join(body_rows), flags_block=flags_block
    )
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w") as f:
        f.write(html)


def main():
    rows = gather_rows()
    rows = add_spreads(rows)
    credit_flags = collect_credit_flags(rows)

    print()
    print_table(rows)
    print_credit_flags(credit_flags)

    xlsx_path = os.path.join(OUTPUT_DIR, "consolidated_comparison.xlsx")
    html_path = os.path.join(OUTPUT_DIR, "consolidated_comparison.html")
    write_xlsx(rows, xlsx_path, credit_flags)
    write_html(rows, html_path, credit_flags)

    print()
    print(f"Wrote {xlsx_path}")
    print(f"Wrote {html_path}")


if __name__ == "__main__":
    main()
